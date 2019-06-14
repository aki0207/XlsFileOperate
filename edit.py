import openpyxl as px
# xlsを対象にしたとき、ファイル形式が古いといわれた
# xlsxなら操作できた

wb = px.load_workbook('test_xlrd.xlsx')

# アクティブシート
ws = wb.active

# シート名取得
# sheets = wb.sheetnames
#print(sheets)


# 書き込み
ws['A1'].value = 'Hello NiceGay'
from datetime import datetime as dt
ws.cell(row=2, column=1).value = dt.now()

# 保存
wb.save('test_xlrd.xlsx')